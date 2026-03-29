"""
ETL MRP -> Template de Campanhas PRs - Rampap

Transforma o export bruto do MRP (Sankhya) para o template de campanhas,
filtrando produtos inativos (Ativo="N") e em ruptura (Em ruptura=1).

Uso:
    python etl_mrp.py --mrp <export_mrp.xlsx> --template <template_campanha.xlsx> --output <saida.xlsx>
"""

import argparse
import io
import os
import sys
from copy import copy
from pathlib import Path

# Fix encoding on Windows console
if sys.platform == "win32":
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")
    sys.stderr = io.TextIOWrapper(sys.stderr.buffer, encoding="utf-8", errors="replace")

try:
    from openpyxl import load_workbook
    from openpyxl.utils import get_column_letter
except ImportError:
    print("openpyxl não encontrado. Instale com: pip install openpyxl")
    sys.exit(1)


# --- Constants ---
HEADER_ROW_MRP = 3          # Row with headers in MRP export
DATA_START_ROW_MRP = 4      # First data row in MRP export
MAX_COL_LETTER = "CK"       # Last column to copy (column 89)
MAX_COL_INDEX = 89           # Number of columns A through CK

# Column indices (1-based) for filtering
COL_ATIVO = 68               # BP = column 68 (Ativo)
COL_EM_RUPTURA = 84          # CF = column 84 (Em ruptura)

SHEET_MRP_ATIVOS = "MRP ativos"
SHEET_IT_PR1 = "IT PR1"


def parse_args():
    parser = argparse.ArgumentParser(description="ETL MRP -> Template de Campanhas PRs")
    parser.add_argument("--mrp", required=True, help="Caminho do export MRP do Sankhya (.xlsx)")
    parser.add_argument("--template", required=True, help="Caminho do template de campanha (.xlsx)")
    parser.add_argument("--output", required=False, help="Caminho do arquivo de saída (.xlsx)")
    return parser.parse_args()


def get_output_path(template_path: str, output_path: str | None) -> Path:
    template = Path(template_path)
    if output_path:
        return Path(output_path)
    return template.parent / f"{template.stem}_ETL{template.suffix}"


def extract_mrp_data(mrp_path: str) -> tuple[list[list], int]:
    """
    Extract data from MRP export, skipping metadata rows and last row.
    Returns (rows, original_count).
    """
    print(f"Abrindo export MRP: {mrp_path}")
    wb = load_workbook(mrp_path, data_only=True)
    ws = wb.active

    all_rows = []
    for row in ws.iter_rows(min_row=DATA_START_ROW_MRP, max_col=MAX_COL_INDEX, values_only=True):
        all_rows.append(list(row))

    wb.close()

    # Remove last row (metadata/totals)
    if all_rows:
        all_rows = all_rows[:-1]

    original_count = len(all_rows)
    print(f"  Linhas extraídas: {original_count}")
    return all_rows, original_count


def filter_ativo_n(rows: list[list]) -> list[list]:
    """Keep only rows where Ativo (col BP, index 67) = 'N'."""
    filtered = [r for r in rows if r[COL_ATIVO - 1] == "N"]
    print(f"  Após filtro Ativo='N': {len(filtered)} linhas (removidas: {len(rows) - len(filtered)})")
    return filtered


def filter_em_ruptura(rows: list[list]) -> list[list]:
    """Keep only rows where Em ruptura (col CF, index 83) = 1."""
    filtered = [r for r in rows if r[COL_EM_RUPTURA - 1] == 1]
    print(f"  Após filtro Em ruptura=1: {len(filtered)} linhas (removidas: {len(rows) - len(filtered)})")
    return filtered


def clear_sheet_data(ws, start_row: int = 2):
    """Clear all data from start_row to end, preserving row 1 (headers)."""
    for row in ws.iter_rows(min_row=start_row, max_col=MAX_COL_INDEX):
        for cell in row:
            cell.value = None


def copy_formatting(source_cell, target_cell):
    """Copy cell formatting from source to target."""
    if source_cell.has_style:
        target_cell.font = copy(source_cell.font)
        target_cell.border = copy(source_cell.border)
        target_cell.fill = copy(source_cell.fill)
        target_cell.number_format = source_cell.number_format
        target_cell.protection = copy(source_cell.protection)
        target_cell.alignment = copy(source_cell.alignment)


def write_rows_to_sheet(ws, rows: list[list], start_row: int = 2):
    """Write rows to sheet starting at start_row, columns A through CK only."""
    for i, row_data in enumerate(rows):
        for j, value in enumerate(row_data):
            if j < MAX_COL_INDEX:
                ws.cell(row=start_row + i, column=j + 1, value=value)


def add_it_pr1_formulas(ws, num_rows: int, start_row: int = 2):
    """
    Add calculated formulas to IT PR1 columns CL-CT.
    These formulas compute 90-day sales averages and increment targets.

    Column mapping (1-based):
      AJ=36 (Vlr mes-1), AF=32 (Vlr mes-2), AB=28 (Vlr mes-3)
      AI=35 (Qtd mes-1), AE=31 (Qtd mes-2), AA=27 (Qtd mes-3)
      CL=90, CM=91, CN=92, CO=93, CP=94, CQ=95, CR=96, CS=97, CT=98
    """
    print(f"  Inserindo formulas nas colunas CL-CT ({num_rows} linhas)...")
    for i in range(num_rows):
        row = start_row + i
        # CL: Venda dos ultimos 90 dias = AJ + AF + AB
        ws.cell(row=row, column=90, value=f"=AJ{row}+AF{row}+AB{row}")
        # CM: Media dos ultimos 90 dias = CL / 3
        ws.cell(row=row, column=91, value=f"=CL{row}/3")
        # CN: Increm (%) esperado — manual, leave empty
        # CO: Meta de Faturamento = CM + (CM * CN)
        ws.cell(row=row, column=93, value=f"=CM{row}+(CM{row}*CN{row})")
        # CP: Increm (R$) esperado = CM * CN
        ws.cell(row=row, column=94, value=f"=CM{row}*CN{row}")
        # CQ: Qtd dos ultimos 90 dias = AI + AE + AA
        ws.cell(row=row, column=95, value=f"=AI{row}+AE{row}+AA{row}")
        # CR: QTD Media dos ultimos 90 dias = CQ / 3
        ws.cell(row=row, column=96, value=f"=CQ{row}/3")
        # CS: Campanha — manual, leave empty
        # CT: Promocao — manual, leave empty


def load_data_into_template(template_path: str, output_path: Path,
                            mrp_ativos_rows: list[list],
                            it_pr1_rows: list[list]):
    """
    Open template, clear target sheets, paste filtered data, add formulas, save.
    """
    print(f"Abrindo template: {template_path}")
    wb = load_workbook(template_path)

    # --- MRP ativos ---
    if SHEET_MRP_ATIVOS not in wb.sheetnames:
        print(f"  ERRO: Aba '{SHEET_MRP_ATIVOS}' nao encontrada no template!")
        sys.exit(1)

    ws_mrp = wb[SHEET_MRP_ATIVOS]
    print(f"  Limpando aba '{SHEET_MRP_ATIVOS}'...")
    clear_sheet_data(ws_mrp, start_row=2)
    print(f"  Escrevendo {len(mrp_ativos_rows)} linhas em '{SHEET_MRP_ATIVOS}'...")
    write_rows_to_sheet(ws_mrp, mrp_ativos_rows, start_row=2)

    # --- IT PR1 ---
    if SHEET_IT_PR1 not in wb.sheetnames:
        print(f"  ERRO: Aba '{SHEET_IT_PR1}' nao encontrada no template!")
        sys.exit(1)

    ws_it = wb[SHEET_IT_PR1]
    print(f"  Limpando aba '{SHEET_IT_PR1}' (colunas A-CK)...")
    clear_sheet_data(ws_it, start_row=2)
    print(f"  Escrevendo {len(it_pr1_rows)} linhas em '{SHEET_IT_PR1}'...")
    write_rows_to_sheet(ws_it, it_pr1_rows, start_row=2)

    # Add formulas to CL-CT
    add_it_pr1_formulas(ws_it, len(it_pr1_rows), start_row=2)

    # --- Save ---
    print(f"Salvando resultado em: {output_path}")
    wb.save(str(output_path))
    wb.close()
    print("ETL concluído com sucesso!")


def main():
    args = parse_args()
    output_path = get_output_path(args.template, args.output)

    print("=" * 60)
    print("ETL MRP -> Template de Campanhas PRs - Rampap")
    print("=" * 60)

    # Extract
    print("\n[1/4] EXTRACT - Extraindo dados do MRP...")
    rows, original_count = extract_mrp_data(args.mrp)

    # Transform
    print("\n[2/4] TRANSFORM - Filtrando Ativo='N'...")
    mrp_ativos_rows = filter_ativo_n(rows)

    print("\n[3/4] TRANSFORM - Filtrando Em ruptura=1...")
    it_pr1_rows = filter_em_ruptura(mrp_ativos_rows)

    # Load
    print("\n[4/4] LOAD - Carregando no template...")
    load_data_into_template(args.template, output_path, mrp_ativos_rows, it_pr1_rows)

    # Summary
    print("\n" + "=" * 60)
    print("RESUMO DO ETL")
    print("=" * 60)
    print(f"  Export MRP original:    {original_count} linhas")
    print(f"  MRP ativos (Ativo=N):   {len(mrp_ativos_rows)} linhas")
    print(f"  IT PR1 (Ruptura=1):     {len(it_pr1_rows)} linhas")
    print(f"  Arquivo gerado:         {output_path}")
    print("=" * 60)


if __name__ == "__main__":
    main()
